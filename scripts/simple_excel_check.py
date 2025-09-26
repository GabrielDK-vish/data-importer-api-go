#!/usr/bin/env python3
"""
Script simples para verificar o arquivo Excel
"""
import openpyxl
import sys
import os

def check_excel(file_path):
    try:
        # Verificar se arquivo existe
        if not os.path.exists(file_path):
            print(f"❌ Arquivo não encontrado: {file_path}")
            return
        
        # Abrir arquivo Excel
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        print(f"📊 Arquivo: {file_path}")
        print(f"📏 Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")
        print(f"📋 Colunas encontradas:")
        
        # Obter cabeçalhos da primeira linha
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Coluna_{col}")
        
        for i, col in enumerate(headers):
            print(f"  {i:2d}: {col}")
        
        # Verificar colunas obrigatórias
        required = ['partner_id', 'customer_id', 'product_id', 'usage_date', 'quantity', 'unit_price']
        print(f"\n🔍 Verificação de colunas obrigatórias:")
        
        missing = []
        for col in required:
            if col in headers:
                print(f"  ✅ {col}")
            else:
                print(f"  ❌ {col} - FALTANDO")
                missing.append(col)
        
        if missing:
            print(f"\n❌ Colunas obrigatórias não encontradas: {missing}")
            print("💡 Sugestões de mapeamento:")
            for col in headers:
                col_lower = col.lower().replace(' ', '').replace('_', '').replace('-', '')
                if 'partner' in col_lower and 'id' in col_lower:
                    print(f"  '{col}' -> partner_id")
                elif 'customer' in col_lower and 'id' in col_lower:
                    print(f"  '{col}' -> customer_id")
                elif 'product' in col_lower and 'id' in col_lower:
                    print(f"  '{col}' -> product_id")
                elif 'usage' in col_lower and 'date' in col_lower:
                    print(f"  '{col}' -> usage_date")
                elif 'quantity' in col_lower:
                    print(f"  '{col}' -> quantity")
                elif 'price' in col_lower and 'unit' in col_lower:
                    print(f"  '{col}' -> unit_price")
        else:
            print(f"\n✅ Todas as colunas obrigatórias foram encontradas!")
        
        # Mostrar primeiras linhas
        print(f"\n📄 Primeiras 3 linhas:")
        for row in range(1, min(4, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(6, sheet.max_column + 1)):  # Mostrar apenas 5 colunas
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell_value)[:20])  # Limitar tamanho
            print(f"  Linha {row}: {' | '.join(row_data)}")
        
    except Exception as e:
        print(f"❌ Erro ao processar arquivo: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        check_excel(sys.argv[1])
    else:
        check_excel("Reconfile fornecedores.xlsx")

